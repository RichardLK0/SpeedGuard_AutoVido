#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
warningSoundDetection_unit
读取配置文件中的 Output 路径，遍历其中的 Excel 文件，根据列数据判断并写入 Sheet2 的 C 列：
- 若 A 列（除表头）无数据或全为空 -> Sheet2中A列值为"S"的行的C列写入 "关闭"
- 若 A 列有数据，且 G 列（除表头）有数据 -> Sheet2中A列值为"S"的行的C列写入 "语音"
- 若 A 列有数据，但 G 列（除表头）无数据 -> Sheet2中A列值为"S"的行的C列写入 "音效"
"""

import sys
import json
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook


def get_script_dir() -> Path:
    """获取脚本所在目录路径。"""
    return Path(__file__).parent.absolute()


def load_output_dir_from_config(config_path: Path) -> Optional[Path]:
    """从 data_process_config.json 中加载 Output 路径。"""
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            config = json.load(f)
        output_path = config.get("output_path")
        if not output_path:
            print("配置文件中缺少 output_path 配置")
            return None
        return Path(output_path)
    except Exception as e:
        print(f"无法加载配置文件: {e}")
        return None


def column_has_data(ws, col_letter: str) -> bool:
    """
    判断指定列（除表头第一行）是否存在非空数据。
    视以下值为“空”：None、空字符串、仅空白字符。
    """
    # 从第二行开始检查
    for cell in ws[f"{col_letter}2":f"{col_letter}{ws.max_row}"]:
        value = cell[0].value
        if value is None:
            continue
        if isinstance(value, str):
            if value.strip() == "":
                continue
            return True
        # 非字符串，且非 None，视为有数据
        return True
    return False


def decide_value_for_result(has_A: bool, has_G: bool) -> str:
    """根据是否有数据的布尔值决定结果内容。"""
    if not has_A:
        return "关闭"
    return "语音" if has_G else "音效"


def find_s_row_in_sheet2(ws) -> Optional[int]:
    """
    在Sheet2中查找A列值为"S"的行号。
    返回行号（从1开始），如果没找到返回None。
    """
    for row in range(1, ws.max_row + 1):
        cell_value = ws[f'A{row}'].value
        if cell_value == 'S':
            return row
    return None


def process_excel_file(xlsx_path: Path) -> None:
    """打开并处理单个 Excel 文件，按规则写入Sheet2的C列和D列公式。"""
    try:
        wb = load_workbook(filename=str(xlsx_path))

        # 检查是否有Sheet2
        if len(wb.worksheets) < 2:
            print(f"文件 {xlsx_path.name} 没有Sheet2，跳过处理")
            return

        ws1 = wb.worksheets[0]  # Sheet1
        ws2 = wb.worksheets[1]  # Sheet2

        # 在Sheet1中检查A列和G列数据
        has_a = column_has_data(ws1, 'A')
        has_g = column_has_data(ws1, 'G')
        result_value = decide_value_for_result(has_a, has_g)

        # 在Sheet2中查找A列值为"S"的行
        s_row = find_s_row_in_sheet2(ws2)
        if s_row is None:
            print(f"在Sheet2中未找到A列值为'S'的行: {xlsx_path.name}")
            return

        # 在找到的行的C列写入结果
        ws2[f'C{s_row}'] = result_value

        # 获取B列的值并比较
        b_value = ws2[f'B{s_row}'].value
        # 处理None值和字符串比较
        if b_value is None:
            b_value = ""
        elif isinstance(b_value, str):
            b_value = b_value.strip()
        else:
            b_value = str(b_value).strip()

        # 比较B列和C列的值
        if b_value == result_value:
            ws2[f'D{s_row}'] = "PASS"
            status_msg = f"PASS (B列='{b_value}' 与 C列相等)"
        else:
            ws2[f'D{s_row}'] = "FAIL"
            status_msg = f"FAIL (B列='{b_value}' 与 C列不相等)"

        print(f"处理完成: {xlsx_path.name} -> Sheet2 C{s_row}='{result_value}', "
              f"D{s_row}='{status_msg}'")

        wb.save(str(xlsx_path))
    except Exception as e:
        print(f"处理文件失败 {xlsx_path}: {e}")


def main() -> None:
    """脚本主入口。"""
    script_dir = get_script_dir()
    config_path = script_dir / 'data_process_config.json'
    output_dir = load_output_dir_from_config(config_path)
    if not output_dir:
        sys.exit(1)
    if not output_dir.exists():
        print(f"输出目录不存在: {output_dir}")
        sys.exit(1)

    excel_files = [
        f for f in output_dir.glob('*.xlsx')
        if not f.name.startswith('~$') and not f.name.startswith('~')
    ]
    if not excel_files:
        print(f"未在输出目录中找到 Excel 文件: {output_dir}")
        sys.exit(0)

    for xlsx in excel_files:
        process_excel_file(xlsx)


if __name__ == '__main__':
    main()
